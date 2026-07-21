# modules/report_generator.py
import os
import csv
import shutil
from datetime import datetime
import pytz
import logging
import time
import threading

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("⚠️ openpyxl не установлен")


class ReportGenerator:
    def __init__(self, file_manager, db):
        self.fm = file_manager
        self.db = db
        self._generating = {}
        self._lock = threading.Lock()
        self._report_downloads = {}

    def generate_report(self, user_id):
        """
        Генерирует отчет ИЗ ТОГО, ЧТО УЖЕ ЕСТЬ в БД.
        НЕ ЖДЕТ все публикации.
        """
        with self._lock:
            if user_id in self._generating:
                elapsed = time.time() - self._generating[user_id]
                if elapsed < 60:
                    logger.warning(f"⚠️ Генерация уже выполняется для {user_id}")
                    return None
            self._generating[user_id] = time.time()
        
        try:
            user_folder = self.fm.get_user_folder(user_id)
            
            # ПОЛУЧАЕМ ВСЕ ПУБЛИКАЦИИ
            publications = self.db.get_publications(user_id)
            
            if not publications:
                logger.warning(f"⚠️ Нет публикаций для {user_id}")
                with self._lock:
                    del self._generating[user_id]
                return None
            
            # Разделяем по статусам
            success_publications = [p for p in publications if p.get('status') == 'success']
            pending_publications = [p for p in publications if p.get('status') == 'pending']
            error_publications = [p for p in publications if p.get('status') != 'success' and p.get('status') != 'pending']
            
            logger.info(f"📊 Статистика: {len(success_publications)} успешных, {len(pending_publications)} ожидают, {len(error_publications)} с ошибками")
            
            # ЕСЛИ НЕТ УСПЕШНЫХ - СОЗДАЕМ ОТЧЕТ С ПРЕДУПРЕЖДЕНИЕМ
            if not success_publications:
                logger.warning(f"⚠️ Нет успешных публикаций для {user_id}")
                return self._create_empty_report(user_id, pending_publications, error_publications)
            
            moscow_tz = pytz.timezone('Europe/Moscow')
            success_data = []
            
            publications_sorted = sorted(success_publications, key=lambda x: x.get('created_at', ''))
            current_date = None
            index = 1
            
            for pub in publications_sorted:
                folder_name = pub.get('folder_name', '')
                chat_id = pub.get('group_id', '')
                
                metadata = self.db.get_ad_metadata(user_id, folder_name)
                
                post_link = metadata.get('post_link', '')
                source_link = metadata.get('Ссылка', '')
                
                if not post_link:
                    post_link = f"https://max.ru/c/{chat_id}" if chat_id else '⚠️ Ссылка не получена'
                
                created_at = pub.get('created_at')
                if created_at:
                    if isinstance(created_at, str):
                        try:
                            created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                        except:
                            created_at = datetime.now(moscow_tz)
                    
                    if hasattr(created_at, 'tzinfo') and created_at.tzinfo is None:
                        created_at = moscow_tz.localize(created_at)
                    elif hasattr(created_at, 'tzinfo'):
                        created_at = created_at.astimezone(moscow_tz)
                    
                    date_str = created_at.strftime('%d.%m.%Y')
                    time_str = created_at.strftime('%H.%M')
                else:
                    now = datetime.now(moscow_tz)
                    date_str = now.strftime('%d.%m.%Y')
                    time_str = now.strftime('%H.%M')
                
                if current_date != date_str:
                    current_date = date_str
                    display_date = date_str
                else:
                    display_date = ''
                
                success_data.append({
                    '№': index,
                    'Дата': display_date,
                    'Время публикации (МСК)': time_str,
                    'Ссылка на пост': post_link,
                    'Ссылка (источник)': source_link,
                    'Название': metadata.get('Название', ''),
                    'Код предложения': metadata.get('Код предложения', ''),
                    'Цена в лизинге': metadata.get('Цена в лизинге', ''),
                })
                index += 1
            
            # ДОБАВЛЯЕМ PENDING И ОШИБКИ В ОТЧЕТ
            if pending_publications:
                success_data.append({})
                success_data.append({
                    '№': '⏳',
                    'Дата': '',
                    'Время публикации (МСК)': '',
                    'Ссылка на пост': f'⏳ {len(pending_publications)} публикаций ожидают ссылки',
                    'Ссылка (источник)': 'Подождите несколько минут',
                    'Название': 'Обновите страницу и проверьте статус',
                    'Код предложения': '',
                    'Цена в лизинге': '',
                })
            
            if error_publications:
                success_data.append({})
                success_data.append({
                    '№': '❌',
                    'Дата': '',
                    'Время публикации (МСК)': '',
                    'Ссылка на пост': f'❌ {len(error_publications)} публикаций с ошибками',
                    'Ссылка (источник)': 'Проверьте логи',
                    'Название': 'Попробуйте загрузить заново',
                    'Код предложения': '',
                    'Цена в лизинге': '',
                })
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            report_filename = f"Отчет_{timestamp}.xlsx"
            report_path = os.path.join(user_folder, report_filename)
            
            if EXCEL_AVAILABLE:
                self._create_excel_report(report_path, success_data)
            else:
                report_filename = f"Отчет_{timestamp}.csv"
                report_path = os.path.join(user_folder, report_filename)
                self._create_csv_report(report_path, success_data)
            
            logger.info(f"📊 Отчет создан: {report_path}")
            
            with self._lock:
                del self._generating[user_id]
            
            return report_path
            
        except Exception as e:
            logger.error(f"❌ Ошибка создания отчета: {e}")
            import traceback
            traceback.print_exc()
            with self._lock:
                if user_id in self._generating:
                    del self._generating[user_id]
            return None

    def _create_empty_report(self, user_id, pending_publications, error_publications):
        """Создает отчет с сообщением, если нет успешных публикаций"""
        try:
            user_folder = self.fm.get_user_folder(user_id)
            
            empty_data = [{
                '№': '⚠️',
                'Дата': '',
                'Время публикации (МСК)': '',
                'Ссылка на пост': 'Нет успешных публикаций',
                'Ссылка (источник)': '',
                'Название': f'{len(pending_publications)} ожидают, {len(error_publications)} с ошибками',
                'Код предложения': 'Попробуйте позже',
                'Цена в лизинге': '',
            }]
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            report_filename = f"Отчет_{timestamp}.xlsx"
            report_path = os.path.join(user_folder, report_filename)
            
            if EXCEL_AVAILABLE:
                self._create_excel_report(report_path, empty_data)
            else:
                report_filename = f"Отчет_{timestamp}.csv"
                report_path = os.path.join(user_folder, report_filename)
                self._create_csv_report(report_path, empty_data)
            
            logger.info(f"📊 Пустой отчет создан: {report_path}")
            return report_path
            
        except Exception as e:
            logger.error(f"❌ Ошибка создания пустого отчета: {e}")
            return None

    def mark_report_downloaded(self, user_id):
        self._report_downloads[user_id] = time.time()
        logger.info(f"📥 Отмечено скачивание отчета для {user_id}")
        
        def cleanup_after_download():
            time.sleep(30)
            logger.info(f"🧹 Автоочистка данных для {user_id} после скачивания отчета")
            self.db.clear_user_data(user_id)
            user_folder = self.fm.get_user_folder(user_id)
            if os.path.exists(user_folder):
                for item in os.listdir(user_folder):
                    item_path = os.path.join(user_folder, item)
                    if os.path.isdir(item_path):
                        shutil.rmtree(item_path)
                    elif not item.startswith('Отчет_'):
                        try:
                            os.remove(item_path)
                        except:
                            pass
            if user_id in self._report_downloads:
                del self._report_downloads[user_id]
        
        threading.Thread(target=cleanup_after_download, daemon=True).start()

    def _create_excel_report(self, filepath, success_data):
        try:
            wb = Workbook()
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            ws = wb.create_sheet("Отчет", 0)
            headers = ['№', 'Дата', 'Время публикации (МСК)', 'Ссылка на пост', 
                      'Ссылка (источник)', 'Название', 'Код предложения', 'Цена в лизинге']
            
            header_font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
            header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            text_font = Font(size=10, name="Calibri")
            text_alignment_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
            text_alignment_center = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin", color="D0D0D0"),
                right=Side(style="thin", color="D0D0D0"),
                top=Side(style="thin", color="D0D0D0"),
                bottom=Side(style="thin", color="D0D0D0")
            )
            link_font = Font(color="0563C1", underline="single", size=10, name="Calibri")
            
            title = ws.cell(row=1, column=1, value="Отчет по публикациям")
            title.font = Font(bold=True, size=16, name="Calibri", color="1A1A2E")
            title.alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells('A1:H1')
            ws.row_dimensions[1].height = 35
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            ws.row_dimensions[2].height = 25
            
            for row_idx, data in enumerate(success_data, 3):
                for col_idx, key in enumerate(headers, 1):
                    value = data.get(key, '')
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    
                    if key in ['№', 'Дата', 'Время публикации (МСК)']:
                        cell.alignment = text_alignment_center
                    elif key in ['Ссылка на пост', 'Ссылка (источник)'] and value and not value.startswith(('⚠️', '⏳', '❌')):
                        cell.font = link_font
                        cell.alignment = text_alignment_left
                    else:
                        cell.font = text_font
                        cell.alignment = text_alignment_left
                    
                    cell.border = thin_border
                    ws.row_dimensions[row_idx].height = 22
            
            for row_idx in range(3, len(success_data) + 3):
                if row_idx % 2 == 1:
                    for col in range(1, 9):
                        cell = ws.cell(row=row_idx, column=col)
                        cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
            
            column_widths = {'A': 5, 'B': 14, 'C': 18, 'D': 50, 'E': 40, 'F': 32, 'G': 18, 'H': 18}
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width
            ws.freeze_panes = 'A3'
            
            wb.save(filepath)
            logger.info(f"✅ Excel отчет создан: {filepath}")
            
        except Exception as e:
            logger.error(f"❌ Ошибка создания Excel: {e}")
            raise

    def _create_csv_report(self, filepath, success_data):
        try:
            with open(filepath, 'w', encoding='utf-8-sig', newline='') as f:
                writer = csv.writer(f, delimiter=';')
                writer.writerow(['№', 'Дата', 'Время публикации (МСК)', 'Ссылка на пост', 
                               'Ссылка (источник)', 'Название', 'Код предложения', 'Цена в лизинге'])
                for data in success_data:
                    writer.writerow([
                        data.get('№', ''),
                        data.get('Дата', ''),
                        data.get('Время публикации (МСК)', ''),
                        data.get('Ссылка на пост', ''),
                        data.get('Ссылка (источник)', ''),
                        data.get('Название', ''),
                        data.get('Код предложения', ''),
                        data.get('Цена в лизинге', ''),
                    ])
            logger.info(f"✅ CSV отчет создан: {filepath}")
        except Exception as e:
            logger.error(f"❌ Ошибка создания CSV: {e}")
            raise

    def cleanup_user_data(self, user_id, keep_report=True):
        try:
            user_folder = self.fm.get_user_folder(user_id)
            if not os.path.exists(user_folder):
                return
            
            if keep_report:
                for item in os.listdir(user_folder):
                    item_path = os.path.join(user_folder, item)
                    if os.path.isdir(item_path):
                        shutil.rmtree(item_path)
                    elif not item.startswith('Отчет_'):
                        try:
                            os.remove(item_path)
                        except:
                            pass
            else:
                shutil.rmtree(user_folder)
                os.makedirs(user_folder, exist_ok=True)
        except Exception as e:
            logger.error(f"❌ Ошибка очистки: {e}")
